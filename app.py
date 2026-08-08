import streamlit as st
import pandas as pd
import numpy as np
import re
import tempfile
from datetime import datetime, date, timezone, timedelta
from dateutil.parser import parse as dt_parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy as pycopy
import hashlib
import time
import os
import warnings
import io

warnings.filterwarnings("ignore")

# =========================
# App config & security
# =========================
APP_VERSION = "3.0.0"
APP_NAME = "Bank Reconciliation"
DEPLOYMENT_MODE = os.environ.get("DEPLOYMENT_MODE", "production")
SESSION_TIMEOUT_MINUTES = 60

st.set_page_config(
    page_title=f"{APP_NAME} v{APP_VERSION}",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

def safe_rerun():
    if hasattr(st, "rerun"):
        st.rerun()
    elif hasattr(st, "experimental_rerun"):
        st.experimental_rerun()

def get_org_password():
    env_pw = os.environ.get("APP_PASSWORD", "").strip()
    if env_pw:
        return env_pw
    try:
        sec_pw = str(st.secrets.get("app_password", "")).strip()
        if sec_pw:
            return sec_pw
    except Exception:
        pass
    return "recon2024"

ORG_PASSWORD = get_org_password()

# =========================
# Theme (Wells Fargo red accent)
# =========================
THEME = {
    "bg": "#ffffff",
    "panel": "#ffffff",
    "panel2": "#f7f7f7",
    "text": "#111111",
    "muted": "#5b5b5b",
    "border": "rgba(0,0,0,0.10)",
    "border2": "rgba(0,0,0,0.14)",
    "accent": "#D71E28",
    "accent2": "#b5161f",
    "good": "#168a45",
    "bad": "#d11a2a",
    "neutral": "#6b7280",
}

def apply_style():
    st.markdown(
        f"""
        <style>
        :root {{
            --bg: {THEME['bg']};
            --panel: {THEME['panel']};
            --panel2: {THEME['panel2']};
            --text: {THEME['text']};
            --muted: {THEME['muted']};
            --border: {THEME['border']};
            --border2: {THEME['border2']};
            --accent: {THEME['accent']};
            --accent2: {THEME['accent2']};
            --good: {THEME['good']};
            --bad: {THEME['bad']};
            --neutral: {THEME['neutral']};
        }}
        html {{
            color-scheme: light !important;
        }}
        html, body, [data-testid="stAppViewContainer"], .stApp {{
            background: var(--bg) !important;
            color: var(--text) !important;
        }}
        [data-testid="stHeader"], [data-testid="stToolbar"], #MainMenu, footer {{
            display: none !important;
        }}
        .block-container {{
            max-width: 1400px;
            padding-top: 2.6rem !important;
            padding-bottom: 2.2rem !important;
        }}
        div[data-testid="stFileUploader"] {{
            background: var(--panel) !important;
            border: 1px dashed var(--border2) !important;
            border-radius: 16px !important;
            padding: 10px !important;
            transition: border 0.2s ease;
        }}
        div[data-testid="stFileUploader"]:hover {{
            border: 1px dashed var(--accent) !important;
        }}
        div[data-testid="stFileUploader"] label {{
            color: var(--text) !important;
            font-weight: 800 !important;
            font-size: 14px !important;
        }}
        div[data-testid="stFileUploader"] small {{
            color: var(--muted) !important;
            font-size: 12px !important;
        }}
        .card {{
            background: #ffffff !important;
            border: 1px solid var(--border) !important;
            border-radius: 18px !important;
            padding: 18px 18px !important;
        }}
        .hero {{
            border: 1px solid var(--border) !important;
            border-radius: 22px !important;
            padding: 26px 22px !important;
            background: radial-gradient(900px 260px at 50% -10%, rgba(215,30,40,0.10), transparent 60%),
                        linear-gradient(180deg, #ffffff, #ffffff) !important;
        }}
        .title {{
            font-size: 30px !important;
            font-weight: 800 !important;
            letter-spacing: 0.2px !important;
            margin: 0 !important;
        }}
        .subtitle {{
            margin-top: 8px !important;
            color: var(--muted) !important;
            font-size: 14px !important;
        }}
        .chip {{
            display: inline-flex !important;
            align-items: center !important;
            gap: 8px !important;
            padding: 6px 12px !important;
            border-radius: 999px !important;
            border: 1px solid var(--border) !important;
            background: #ffffff !important;
            font-size: 12px !important;
            font-weight: 650 !important;
            color: var(--muted) !important;
        }}
        .chip-dot {{
            width: 8px !important;
            height: 8px !important;
            border-radius: 999px !important;
            display: inline-block !important;
            background: var(--accent) !important;
        }}
        .metric {{
            border: 1px solid var(--border) !important;
            border-radius: 18px !important;
            padding: 14px 14px !important;
            background: #ffffff !important;
        }}
        .metric-k {{
            font-size: 12px !important;
            color: var(--muted) !important;
            font-weight: 700 !important;
            text-transform: uppercase !important;
        }}
        .metric-v {{
            font-size: 26px !important;
            font-weight: 850 !important;
            margin-top: 6px !important;
            color: var(--accent) !important;
        }}
        div.stButton > button {{
            background: var(--accent) !important;
            border: 1px solid var(--accent) !important;
            border-radius: 14px !important;
            padding: 0.7rem 1rem !important;
            font-weight: 750 !important;
            color: #ffffff !important;
        }}
        div.stButton > button:hover {{
            background: var(--accent2) !important;
            border-color: var(--accent2) !important;
        }}
        div[data-baseweb="base-input"] > div,
        div[data-baseweb="input"] > div {{
            background: #ffffff !important;
            border: 1px solid var(--border2) !important;
            border-radius: 14px !important;
        }}
        [data-testid="stDataFrame"] {{
            background: #ffffff !important;
            border: 1px solid var(--border) !important;
            border-radius: 16px !important;
            overflow: hidden !important;
        }}
        a, a:visited {{
            color: var(--accent) !important;
            font-weight: 750 !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

apply_style()

# =========================
# Session management (authentication)
# =========================
def touch():
    st.session_state.last_activity = datetime.now()

def is_timed_out():
    last = st.session_state.get("last_activity")
    if not last:
        return False
    return (datetime.now() - last).total_seconds() > SESSION_TIMEOUT_MINUTES * 60

def logout():
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    safe_rerun()

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "session_id" not in st.session_state:
    st.session_state.session_id = hashlib.sha256(str(time.time()).encode()).hexdigest()[:16]
if "last_activity" not in st.session_state:
    st.session_state.last_activity = datetime.now()

def login_screen():
    st.markdown('<div style="height: 1.8rem;"></div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1.25, 1])
    with c2:
        st.markdown(
            f"""
            <div class="card" style="margin-top: 10vh;">
                <div class="title" style="text-align:center;">{APP_NAME}</div>
                <div class="subtitle" style="text-align:center;">Sign in to continue.</div>
                <div style="height: 14px;"></div>
                <div style="display:flex; justify-content:center;">
                    <div class="chip"><span class="chip-dot"></span> Version {APP_VERSION} • {DEPLOYMENT_MODE.title()}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.form("login_form", clear_on_submit=True):
            pw = st.text_input("Password", type="password", placeholder="Organisation password")
            ok = st.form_submit_button("Sign in", use_container_width=True)
        if ok:
            if pw == ORG_PASSWORD:
                st.session_state.authenticated = True
                touch()
                safe_rerun()
            else:
                st.error("Wrong password.")

if st.session_state.authenticated and is_timed_out():
    st.session_state.authenticated = False
    st.warning("Session timed out. Sign in again.")
    login_screen()
    st.stop()

if not st.session_state.authenticated:
    login_screen()
    st.stop()

touch()

# =========================
# Session state for reconciliation results
# =========================
def init_session_state():
    if "reconciliation_done" not in st.session_state:
        st.session_state.reconciliation_done = False
    if "bank_df" not in st.session_state:
        st.session_state.bank_df = None
    if "ledger_df" not in st.session_state:
        st.session_state.ledger_df = None
    if "match_results" not in st.session_state:
        st.session_state.match_results = None
    if "working_paper" not in st.session_state:
        st.session_state.working_paper = None
    if "recon_statement" not in st.session_state:
        st.session_state.recon_statement = None
    if "output_bytes" not in st.session_state:
        st.session_state.output_bytes = None
    if "output_filename" not in st.session_state:
        st.session_state.output_filename = None
    if "file_info" not in st.session_state:
        st.session_state.file_info = {"bank": None, "ledger": None}
    if "opening_balance_manual" not in st.session_state:
        st.session_state.opening_balance_manual = 0.0

def clear_reconciliation_state():
    st.session_state.reconciliation_done = False
    st.session_state.bank_df = None
    st.session_state.ledger_df = None
    st.session_state.match_results = None
    st.session_state.working_paper = None
    st.session_state.recon_statement = None
    st.session_state.output_bytes = None
    st.session_state.output_filename = None
    st.session_state.file_info = {"bank": None, "ledger": None}

init_session_state()

# =========================
# Helper functions
# =========================
def to_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def to_num(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = to_str(x).replace(",", "")
    if '(' in s and ')' in s:
        s = '-' + s.replace('(', '').replace(')', '').strip()
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s or s == '-':
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan

def excel_serial_to_date(serial):
    if pd.isna(serial) or not isinstance(serial, (int, float)):
        return pd.NaT
    try:
        base = datetime(1899, 12, 30)
        delta = timedelta(days=float(serial))
        return base + delta
    except:
        return pd.NaT

def to_date(x):
    if isinstance(x, (pd.Timestamp, np.datetime64, datetime, date)):
        return pd.to_datetime(x, errors='coerce')
    if isinstance(x, (int, float, np.number)):
        if x > 10000:
            return excel_serial_to_date(x)
        else:
            try:
                return pd.to_datetime(x, unit='s', errors='coerce')
            except:
                return pd.NaT
    s = to_str(x)
    if not s:
        return pd.NaT
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y', '%d-%m-%y']:
        try:
            return pd.to_datetime(s, format=fmt)
        except:
            continue
    try:
        return pd.to_datetime(dt_parse(s, fuzzy=True))
    except Exception:
        return pd.NaT

def format_date(date_val):
    if date_val is None:
        return ''
    if pd.isna(date_val):
        return ''
    if isinstance(date_val, str):
        if date_val == '':
            return ''
        try:
            parsed = pd.to_datetime(date_val, errors='coerce')
            if pd.notna(parsed):
                return parsed.strftime('%d/%m/%y')
            else:
                return date_val[:10] if len(date_val) > 10 else date_val
        except:
            return date_val[:10] if len(date_val) > 10 else date_val
    if isinstance(date_val, (pd.Timestamp, datetime, date)):
        if pd.isna(date_val):
            return ''
        try:
            return date_val.strftime('%d/%m/%y')
        except:
            return ''
    return str(date_val)

def detect_table(df_raw, max_rows=100):
    if df_raw is None or df_raw.empty:
        return None, 0
    best_row = 0
    best_score = 0
    for r in range(min(max_rows, len(df_raw))):
        row = df_raw.iloc[r]
        non_empty = row.notna().sum()
        if non_empty < 2:
            continue
        header_score = 0
        for cell in row:
            if pd.isna(cell):
                continue
            s = str(cell).lower()
            if any(kw in s for kw in ['date', 'amount', 'credit', 'debit', 'ref', 'description', 'posting', 'document', 'narration']):
                header_score += 2
        score = non_empty + header_score
        if score > best_score:
            best_score = score
            best_row = r
    if best_row >= 0 and best_row < len(df_raw):
        headers = df_raw.iloc[best_row].fillna('').astype(str).tolist()
        headers = [h.strip() if h.strip() else f"Column_{i}" for i, h in enumerate(headers)]
        data = df_raw.iloc[best_row + 1:].reset_index(drop=True)
        if data.shape[1] < len(headers):
            for _ in range(len(headers) - data.shape[1]):
                data[f"extra_{_}"] = None
        elif data.shape[1] > len(headers):
            data = data.iloc[:, :len(headers)]
        data.columns = headers
        data = data.dropna(how='all')
        return data, best_row
    return df_raw, 0

def load_bank_statement(file):
    try:
        excel_file = pd.ExcelFile(file)
        sheet_names = excel_file.sheet_names
        visible_sheets = [s for s in sheet_names if not s.startswith('_') and 'hidden' not in s.lower()]
        if not visible_sheets:
            visible_sheets = sheet_names
        df_bank = None
        header_row_used = 0
        for sheet in visible_sheets:
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet, header=None)
                if df_raw is not None and not df_raw.empty:
                    df, header_row = detect_table(df_raw)
                    if df is not None and len(df) > 3:
                        df_bank = df
                        header_row_used = header_row
                        break
            except Exception:
                continue
        if df_bank is None or df_bank.empty:
            df_bank = pd.read_excel(file)
            if df_bank.empty:
                raise ValueError("Could not read bank statement file")
    except Exception:
        df_bank = pd.read_excel(file)

    date_col = None
    credit_col = None
    debit_col = None
    ref_col = None
    desc_col = None
    balance_col = None
    for col in df_bank.columns:
        cl = str(col).lower().strip()
        if 'date' in cl and not date_col:
            date_col = col
        if 'credit' in cl and not credit_col:
            credit_col = col
        if 'debit' in cl and not debit_col:
            debit_col = col
        if 'ref' in cl or 'reference' in cl or 'cheque' in cl:
            ref_col = col
        if 'narrative' in cl or 'description' in cl or 'particulars' in cl or 'transaction' in cl:
            desc_col = col
        if 'balance' in cl and not balance_col:
            balance_col = col
    if not credit_col and not debit_col:
        for col in df_bank.columns:
            if 'amount' in str(col).lower() or 'value' in str(col).lower():
                credit_col = col
                debit_col = col
                break

    opening_balance = None
    for idx, row in df_bank.iterrows():
        row_str = ' '.join(str(v).lower() for v in row.values if pd.notna(v))
        if 'balance at period start' in row_str or 'opening balance' in row_str or 'balance brought forward' in row_str:
            for col in df_bank.columns:
                val = to_num(row[col])
                if not pd.isna(val) and val != 0:
                    opening_balance = val
                    break
            break

    transactions = []
    for idx, row in df_bank.iterrows():
        credit = 0
        debit = 0
        if credit_col and credit_col in df_bank.columns:
            credit = to_num(row[credit_col]) if pd.notna(row[credit_col]) else 0
        if debit_col and debit_col in df_bank.columns:
            debit = to_num(row[debit_col]) if pd.notna(row[debit_col]) else 0
        if credit == 0 and debit == 0 and credit_col == debit_col and credit_col:
            amount = to_num(row[credit_col]) if pd.notna(row[credit_col]) else 0
            if amount == 0:
                continue
            credit = amount if amount > 0 else 0
            debit = -amount if amount < 0 else 0
        if credit == 0 and debit == 0:
            continue
        trans_date = to_date(row[date_col]) if date_col and date_col in df_bank.columns else pd.NaT
        if pd.isna(trans_date):
            continue
        desc = ''
        if desc_col and desc_col in df_bank.columns:
            desc = to_str(row[desc_col])
        if not desc and ref_col and ref_col in df_bank.columns:
            desc = to_str(row[ref_col])
        transactions.append({
            'date': trans_date,
            'reference': to_str(row[ref_col]) if ref_col and ref_col in df_bank.columns else '',
            'description': desc,
            'credit': credit,
            'debit': debit,
            'amount': credit - debit,
            'abs_amount': abs(credit - debit),
            'source': 'BANK'
        })
    df_bank_norm = pd.DataFrame(transactions)
    if df_bank_norm.empty:
        st.error("No valid bank transactions.")
        st.stop()
    closing_balance = None
    if balance_col and balance_col in df_bank.columns:
        last_valid = df_bank[balance_col].dropna().iloc[-1] if not df_bank[balance_col].dropna().empty else None
        closing_balance = to_num(last_valid) if last_valid else None
    return df_bank_norm, opening_balance, closing_balance, header_row_used

def load_ledger(file):
    try:
        excel_file = pd.ExcelFile(file)
        sheet_names = excel_file.sheet_names
        visible_sheets = [s for s in sheet_names if not s.startswith('_') and 'hidden' not in s.lower()]
        if not visible_sheets:
            visible_sheets = sheet_names
        df_ledger = None
        header_row_used = 0
        for sheet in visible_sheets:
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet, header=None)
                if df_raw is not None and not df_raw.empty:
                    df, header_row = detect_table(df_raw)
                    if df is not None and len(df) > 3:
                        df_ledger = df
                        header_row_used = header_row
                        break
            except Exception:
                continue
        if df_ledger is None or df_ledger.empty:
            df_ledger = pd.read_excel(file)
            if df_ledger.empty:
                raise ValueError("Could not read ledger file")
    except Exception:
        df_ledger = pd.read_excel(file)

    date_col = None
    amount_col = None
    desc_col = None
    ref_col = None
    balance_col = None
    for col in df_ledger.columns:
        cl = str(col).lower().strip()
        if 'date' in cl and not date_col:
            date_col = col
        if 'amount' in cl and not amount_col:
            amount_col = col
        if 'description' in cl or 'desc' in cl or 'particulars' in cl or 'details' in cl:
            desc_col = col
        if 'document' in cl or 'ref' in cl or 'external' in cl or 'reference' in cl:
            ref_col = col
        if 'balance' in cl and not balance_col:
            balance_col = col

    transactions = []
    closing_balance = None
    for idx, row in df_ledger.iterrows():
        amount = to_num(row[amount_col]) if amount_col and amount_col in df_ledger.columns else np.nan
        if pd.isna(amount):
            continue
        trans_date = to_date(row[date_col]) if date_col and date_col in df_ledger.columns else pd.NaT
        if pd.isna(trans_date):
            continue
        # Store the running balance if available
        if balance_col and balance_col in df_ledger.columns:
            running_balance = to_num(row[balance_col])
            if not pd.isna(running_balance):
                closing_balance = running_balance  # update to latest non-NaN balance
        transactions.append({
            'date': trans_date,
            'reference': to_str(row[ref_col]) if ref_col and ref_col in df_ledger.columns else '',
            'description': to_str(row[desc_col]) if desc_col and desc_col in df_ledger.columns else '',
            'amount': amount,
            'abs_amount': abs(amount),
            'type': 'CREDIT' if amount > 0 else 'DEBIT',
            'source': 'LEDGER'
        })
    df_ledger_norm = pd.DataFrame(transactions)
    if df_ledger_norm.empty:
        st.error("No valid ledger transactions.")
        st.stop()
    # If a balance column was found, use the last non-null balance as closing balance.
    # Otherwise fall back to the sum of amounts (net movement).
    if closing_balance is None:
        closing_balance = df_ledger_norm['amount'].sum() if not df_ledger_norm.empty else 0
    return df_ledger_norm, closing_balance, header_row_used

# =========================
# Improved Matching Logic
# =========================
def match_transactions(bank_df, ledger_df):
    bank_copy = bank_df.copy()
    ledger_copy = ledger_df.copy()

    bank_credits = bank_copy[bank_copy['credit'] > 0].copy() if 'credit' in bank_copy.columns else bank_copy[bank_copy['amount'] > 0].copy()
    bank_debits = bank_copy[bank_copy['debit'] > 0].copy() if 'debit' in bank_copy.columns else bank_copy[bank_copy['amount'] < 0].copy()
    if 'credit' not in bank_copy.columns:
        bank_credits = bank_copy[bank_copy['amount'] > 0].copy()
        bank_debits = bank_copy[bank_copy['amount'] < 0].copy()

    ledger_credits = ledger_copy[ledger_copy['amount'] > 0].copy()
    ledger_debits = ledger_copy[ledger_copy['amount'] < 0].copy()

    bank_credits['abs_amount_rounded'] = bank_credits['abs_amount'].round(2)
    bank_debits['abs_amount_rounded'] = bank_debits['abs_amount'].round(2)
    ledger_credits['abs_amount_rounded'] = ledger_credits['abs_amount'].round(2)
    ledger_debits['abs_amount_rounded'] = ledger_debits['abs_amount'].round(2)

    for df in [bank_credits, bank_debits, ledger_credits, ledger_debits]:
        if not df.empty:
            df['date'] = pd.to_datetime(df['date'])

    bank_credits['id'] = [f'B_C_{i}' for i in range(len(bank_credits))]
    bank_debits['id'] = [f'B_D_{i}' for i in range(len(bank_debits))]
    ledger_credits['id'] = [f'L_C_{i}' for i in range(len(ledger_credits))]
    ledger_debits['id'] = [f'L_D_{i}' for i in range(len(ledger_debits))]

    matches = []

    def match_group(ledger_items, bank_items, txn_type):
        if ledger_items.empty or bank_items.empty:
            return [], [], []

        led_dict = ledger_items.to_dict('records')
        bank_dict = bank_items.to_dict('records')

        matched_ledger_ids = set()
        matched_bank_ids = set()
        group_matches = []

        # First pass: amount + date
        for l_item in led_dict:
            if l_item['id'] in matched_ledger_ids:
                continue
            for b_item in bank_dict:
                if b_item['id'] in matched_bank_ids:
                    continue
                if abs(l_item['abs_amount_rounded'] - b_item['abs_amount_rounded']) <= 0.01:
                    if l_item['date'].date() == b_item['date'].date():
                        group_matches.append({
                            'ledger_id': l_item['id'],
                            'bank_id': b_item['id'],
                            'amount': l_item['abs_amount_rounded'],
                            'type': txn_type,
                            'match_method': 'amount_and_date',
                            'ledger_date': l_item['date'],
                            'bank_date': b_item['date'],
                            'date_match': True,
                            'ledger_desc': l_item.get('description', ''),
                            'ledger_ref': l_item.get('reference', ''),
                            'bank_desc': b_item.get('description', ''),
                            'bank_ref': b_item.get('reference', '')
                        })
                        matched_ledger_ids.add(l_item['id'])
                        matched_bank_ids.add(b_item['id'])
                        break

        # Second pass: amount only
        for l_item in led_dict:
            if l_item['id'] in matched_ledger_ids:
                continue
            for b_item in bank_dict:
                if b_item['id'] in matched_bank_ids:
                    continue
                if abs(l_item['abs_amount_rounded'] - b_item['abs_amount_rounded']) <= 0.01:
                    group_matches.append({
                        'ledger_id': l_item['id'],
                        'bank_id': b_item['id'],
                        'amount': l_item['abs_amount_rounded'],
                        'type': txn_type,
                        'match_method': 'amount_only',
                        'ledger_date': l_item['date'],
                        'bank_date': b_item['date'],
                        'date_match': False,
                        'ledger_desc': l_item.get('description', ''),
                        'ledger_ref': l_item.get('reference', ''),
                        'bank_desc': b_item.get('description', ''),
                        'bank_ref': b_item.get('reference', '')
                    })
                    matched_ledger_ids.add(l_item['id'])
                    matched_bank_ids.add(b_item['id'])
                    break

        unmatch_ledger = [item for item in led_dict if item['id'] not in matched_ledger_ids]
        unmatch_bank = [item for item in bank_dict if item['id'] not in matched_bank_ids]

        return group_matches, unmatch_ledger, unmatch_bank

    credit_matches, unmatched_ledger_credits, unmatched_bank_credits = match_group(
        ledger_credits, bank_credits, 'CREDIT'
    )
    debit_matches, unmatched_ledger_debits, unmatched_bank_debits = match_group(
        ledger_debits, bank_debits, 'DEBIT'
    )

    matches = credit_matches + debit_matches

    def clean_item(item):
        return {
            'date': item['date'],
            'reference': item.get('reference', ''),
            'description': item.get('description', ''),
            'amount': item['amount'],
            'abs_amount': item.get('abs_amount', abs(item['amount'])),
            'type': item.get('type', 'CREDIT' if item['amount'] > 0 else 'DEBIT'),
            'source': item.get('source', 'LEDGER' if 'id' in item and item['id'].startswith('L') else 'BANK')
        }

    return {
        'matches': matches,
        'unmatched_ledger_credits': [clean_item(x) for x in unmatched_ledger_credits],
        'unmatched_ledger_debits': [clean_item(x) for x in unmatched_ledger_debits],
        'unmatched_bank_credits': [clean_item(x) for x in unmatched_bank_credits],
        'unmatched_bank_debits': [clean_item(x) for x in unmatched_bank_debits]
    }

# =========================
# Build Working Paper
# =========================
def build_working_paper(match_results):
    rows = []
    for match in match_results['matches']:
        rows.append({
            'SECTION': 'MATCHED',
            'LEDGER_DATE': match['ledger_date'],
            'LEDGER_DESC': match.get('ledger_desc',''),
            'LEDGER_REF': match.get('ledger_ref',''),
            'LEDGER_AMOUNT': match['amount'] if match['type']=='CREDIT' else -match['amount'],
            'MATCH_STATUS': f"MATCHED - {match['match_method'].replace('_',' ').upper()}",
            'BANK_AMOUNT': match['amount'] if match['type']=='CREDIT' else -match['amount'],
            'BANK_DATE': match['bank_date'],
            'BANK_REF': match.get('bank_ref',''),
            'BANK_DESC': match.get('bank_desc','')
        })
    for item in match_results['unmatched_ledger_credits']:
        rows.append({
            'SECTION': 'UNMATCHED - LEDGER ONLY',
            'LEDGER_DATE': item['date'], 'LEDGER_DESC': item['description'], 'LEDGER_REF': item['reference'],
            'LEDGER_AMOUNT': item['amount'], 'MATCH_STATUS': 'NO BANK MATCH',
            'BANK_AMOUNT': '', 'BANK_DATE': '', 'BANK_REF': '', 'BANK_DESC': ''
        })
    for item in match_results['unmatched_ledger_debits']:
        rows.append({
            'SECTION': 'UNMATCHED - LEDGER ONLY',
            'LEDGER_DATE': item['date'], 'LEDGER_DESC': item['description'], 'LEDGER_REF': item['reference'],
            'LEDGER_AMOUNT': item['amount'], 'MATCH_STATUS': 'NO BANK MATCH',
            'BANK_AMOUNT': '', 'BANK_DATE': '', 'BANK_REF': '', 'BANK_DESC': ''
        })
    for item in match_results['unmatched_bank_credits']:
        rows.append({
            'SECTION': 'UNMATCHED - BANK ONLY',
            'LEDGER_DATE': '', 'LEDGER_DESC': '', 'LEDGER_REF': '', 'LEDGER_AMOUNT': '',
            'MATCH_STATUS': 'NO LEDGER MATCH',
            'BANK_AMOUNT': item['amount'], 'BANK_DATE': item['date'], 'BANK_REF': item['reference'], 'BANK_DESC': item['description']
        })
    for item in match_results['unmatched_bank_debits']:
        rows.append({
            'SECTION': 'UNMATCHED - BANK ONLY',
            'LEDGER_DATE': '', 'LEDGER_DESC': '', 'LEDGER_REF': '', 'LEDGER_AMOUNT': '',
            'MATCH_STATUS': 'NO LEDGER MATCH',
            'BANK_AMOUNT': item['amount'], 'BANK_DATE': item['date'], 'BANK_REF': item['reference'], 'BANK_DESC': item['description']
        })
    return pd.DataFrame(rows)

# =========================
# Build Reconciliation Statement
# =========================
def build_recon_statement(bank_opening, bank_closing, ledger_closing, match_results):
    recon_items = []
    total_adjustment = 0.0

    for item in match_results['unmatched_bank_credits']:
        amt = -abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"BANK ONLY - CREDIT TO POST: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Bank-only credit'
        })
        total_adjustment += amt

    for item in match_results['unmatched_bank_debits']:
        amt = abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"BANK ONLY - DEBIT TO POST: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Bank-only debit'
        })
        total_adjustment += amt

    for item in match_results['unmatched_ledger_credits']:
        amt = abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"LEDGER ONLY - DEPOSIT IN TRANSIT: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Ledger-only credit'
        })
        total_adjustment += amt

    for item in match_results['unmatched_ledger_debits']:
        amt = -abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"LEDGER ONLY - UNPRESENTED PAYMENT: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Ledger-only debit'
        })
        total_adjustment += amt

    bank_closing = float(bank_closing or 0)
    ledger_closing = float(ledger_closing or 0)
    adjusted_balance = bank_closing + total_adjustment
    difference = adjusted_balance - ledger_closing

    bank_movement = bank_closing - float(bank_opening or 0)
    ledger_movement = ledger_closing - float(bank_opening or 0)

    return {
        'opening_balance': float(bank_opening or 0),
        'bank_closing_balance': bank_closing,
        'recon_items': pd.DataFrame(recon_items),
        'total_adjustment': total_adjustment,
        'adjusted_balance': adjusted_balance,
        'ledger_balance': ledger_closing,
        'difference': difference,
        'bank_movement': bank_movement,
        'ledger_movement': ledger_movement,
        'movement_difference': bank_movement - ledger_movement
    }

# =========================
# Export to Excel
# =========================
def export_to_excel(working_paper_df, recon_statement, bank_df, ledger_df, match_results):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        wp_display = working_paper_df.copy()
        wp_display['LEDGER_DATE'] = wp_display['LEDGER_DATE'].apply(lambda x: format_date(x) if pd.notna(x) else '')
        wp_display['BANK_DATE'] = wp_display['BANK_DATE'].apply(lambda x: format_date(x) if pd.notna(x) else '')
        wp_display['LEDGER_AMOUNT'] = wp_display['LEDGER_AMOUNT'].apply(
            lambda x: f"{float(x):,.2f}" if pd.notna(x) and x != '' else ''
        )
        wp_display['BANK_AMOUNT'] = wp_display['BANK_AMOUNT'].apply(
            lambda x: f"{float(x):,.2f}" if pd.notna(x) and x != '' else ''
        )
        wp_display.to_excel(writer, sheet_name='WORKING_PAPER', index=False)

        recon_data = [
            ['BANK RECONCILIATION STATEMENT', '', ''],
            ['', '', ''],
            ['Opening bank balance', '', f"{recon_statement['opening_balance']:,.2f}"],
            ['Bank closing balance per statement', '', f"{recon_statement['bank_closing_balance']:,.2f}"],
            ['', '', ''],
            ['RECONCILING ITEMS', '', ''],
            ['Date', 'Description', 'Adjustment',],
        ]

        for _, item in recon_statement['recon_items'].iterrows():
            recon_data.append([
                format_date(item['date']) if pd.notna(item['date']) else '',
                item['description'][:120],
                f"{item['adjustment']:,.2f}"
            ])

        recon_data += [
            ['', 'Total reconciling adjustments', f"{recon_statement['total_adjustment']:,.2f}"],
            ['', 'Adjusted bank balance', f"{recon_statement['adjusted_balance']:,.2f}"],
            ['', 'Ledger closing balance', f"{recon_statement['ledger_balance']:,.2f}"],
            ['', 'UNRECONCILED DIFFERENCE', f"{recon_statement['difference']:,.2f}"],
            ['', '', ''],
            ['DIAGNOSTIC ANALYSIS', '', ''],
            ['Bank movement during period', f"{recon_statement['bank_movement']:,.2f}", ''],
            ['Ledger movement using same opening basis', f"{recon_statement['ledger_movement']:,.2f}", ''],
            ['Movement difference', f"{recon_statement['movement_difference']:,.2f}", ''],
            ['Opening balance basis', f"{recon_statement['opening_balance']:,.2f}", ''],
            ['', '', ''],
            ['Interpretation', 'Difference must be investigated; do not post a balancing adjustment automatically.', ''],
            ['', '', ''],
            ['Prepared by:', '', ''],
            ['Date:', '', datetime.now().strftime('%d/%m/%y')]
        ]
        pd.DataFrame(recon_data).to_excel(writer, sheet_name='RECON_STATEMENT', index=False, header=False)

        matched_detail = [['Matched Transactions - Detailed View'], ['']]
        matched_detail.append([
            'Ledger Date','Ledger Description','Ledger Ref','Ledger Amount',
            'Bank Date','Bank Ref','Bank Description','Bank Amount','Match Method'
        ])
        for match in match_results['matches']:
            signed_amount = match['amount'] if match['type'] == 'CREDIT' else -match['amount']
            matched_detail.append([
                format_date(match['ledger_date']),
                (match.get('ledger_desc','')[:80] if match.get('ledger_desc') else ''),
                (match.get('ledger_ref','')[:40] if match.get('ledger_ref') else ''),
                f"{signed_amount:,.2f}",
                format_date(match['bank_date']),
                (match.get('bank_ref','')[:40] if match.get('bank_ref') else ''),
                (match.get('bank_desc','')[:80] if match.get('bank_desc') else ''),
                f"{signed_amount:,.2f}",
                match['match_method']
            ])
        pd.DataFrame(matched_detail).to_excel(writer, sheet_name='MATCHED_DETAIL', index=False, header=False)

        uml = [['LEDGER TRANSACTIONS WITH NO BANK MATCH'], [''],
               ['Date','Description','Reference','Amount','Type']]
        for item in match_results['unmatched_ledger_credits']:
            uml.append([format_date(item['date']), item['description'][:80], item['reference'][:40],
                        f"{item['amount']:,.2f}", 'CREDIT / DEPOSIT IN TRANSIT'])
        for item in match_results['unmatched_ledger_debits']:
            uml.append([format_date(item['date']), item['description'][:80], item['reference'][:40],
                        f"{item['amount']:,.2f}", 'DEBIT / UNPRESENTED PAYMENT'])
        pd.DataFrame(uml).to_excel(writer, sheet_name='UNMATCHED_LEDGER', index=False, header=False)

        umb = [['BANK TRANSACTIONS WITH NO LEDGER MATCH'], [''],
               ['Date','Description','Reference','Amount','Type']]
        for item in match_results['unmatched_bank_credits']:
            umb.append([format_date(item['date']), item['description'][:80], item['reference'][:40],
                        f"{item['amount']:,.2f}", 'CREDIT / POST TO LEDGER'])
        for item in match_results['unmatched_bank_debits']:
            umb.append([format_date(item['date']), item['description'][:80], item['reference'][:40],
                        f"{item['amount']:,.2f}", 'DEBIT / POST TO LEDGER'])
        pd.DataFrame(umb).to_excel(writer, sheet_name='UNMATCHED_BANK', index=False, header=False)

        summary_data = [
            ['RECONCILIATION SUMMARY', ''],
            ['', ''],
            ['Total bank transactions', len(bank_df)],
            ['Total ledger transactions', len(ledger_df)],
            ['Matched transactions', len(match_results['matches'])],
            ['Unmatched - ledger only',
             len(match_results['unmatched_ledger_credits']) + len(match_results['unmatched_ledger_debits'])],
            ['Unmatched - bank only',
             len(match_results['unmatched_bank_credits']) + len(match_results['unmatched_bank_debits'])],
            ['', ''],
            ['Opening bank balance', f"{recon_statement['opening_balance']:,.2f}"],
            ['Bank closing balance', f"{recon_statement['bank_closing_balance']:,.2f}"],
            ['Total reconciling adjustments', f"{recon_statement['total_adjustment']:,.2f}"],
            ['Adjusted bank balance', f"{recon_statement['adjusted_balance']:,.2f}"],
            ['Ledger closing balance', f"{recon_statement['ledger_balance']:,.2f}"],
            ['Unreconciled difference', f"{recon_statement['difference']:,.2f}"],
            ['', ''],
            ['Status', 'RECONCILED' if abs(recon_statement['difference']) < 0.01 else 'INVESTIGATION REQUIRED']
        ]
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='SUMMARY', index=False, header=False)

        for sheetname, ws in writer.sheets.items():
            ws.freeze_panes = 'A2'
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    return output.getvalue()

# =========================
# Main UI – with personalized header
# =========================
st.markdown(
    f"""
    <div class="hero" style="text-align:center;">
        <div class="title">🏦 Bank Reconciliation <span style="color: {THEME['accent']};">with Chipo</span></div>
        <div class="subtitle">Welcome, Chipo! Upload your Bank Statement and Ledger to reconcile closing balances, explain differences, and download a detailed working paper.</div>
        <div style="height: 12px;"></div>
        <div style="display:flex; justify-content:center; gap:10px; flex-wrap:wrap;">
            <div class="chip"><span class="chip-dot"></span> Secure session</div>
            <div class="chip">Session {st.session_state.session_id}</div>
            <div class="chip">Mode {DEPLOYMENT_MODE.title()}</div>
            <div class="chip">Version {APP_VERSION}</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
st.markdown("")

def files_changed(bank_file, ledger_file):
    current_bank = (bank_file.name, bank_file.size) if bank_file else None
    current_ledger = (ledger_file.name, ledger_file.size) if ledger_file else None
    old_bank = st.session_state.file_info.get("bank")
    old_ledger = st.session_state.file_info.get("ledger")
    return (current_bank != old_bank) or (current_ledger != old_ledger)

col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("🏦 Bank Statement")
    bank_file = st.file_uploader(
        "Bank Statement (Excel file)",
        type=["xlsx", "xls"],
        help="Upload your bank statement Excel file. Required columns: Date, Credit, Debit, Description (or similar).",
        key="bank"
    )
    st.caption("Accepts .xlsx or .xls, max 200MB per file")
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("📒 Cashbook / Ledger")
    ledger_file = st.file_uploader(
        "Ledger / Cashbook (Excel file)",
        type=["xlsx", "xls"],
        help="Upload your Yellowcob ledger Excel file. Required columns: Date, Amount, Description (or similar).",
        key="ledger"
    )
    st.caption("Accepts .xlsx or .xls, max 200MB per file")
    st.markdown('</div>', unsafe_allow_html=True)

if bank_file and ledger_file:
    if files_changed(bank_file, ledger_file):
        clear_reconciliation_state()
        st.session_state.file_info["bank"] = (bank_file.name, bank_file.size)
        st.session_state.file_info["ledger"] = (ledger_file.name, ledger_file.size)

st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("⚙️ Reconciliation Settings")
st.markdown("Enter the Bank Opening Balance below. If you leave it as 0, the app will try to auto-detect it from the bank statement.")
opening_balance_manual = st.number_input(
    "Bank Opening Balance (manual override)",
    value=st.session_state.opening_balance_manual,
    step=100.00,
    format="%.2f",
    key="opening_balance_input"
)
st.session_state.opening_balance_manual = opening_balance_manual
st.markdown('</div>', unsafe_allow_html=True)

col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
with col_btn1:
    run_recon = st.button("🔄 RUN RECONCILIATION", use_container_width=True)
with col_btn2:
    if st.button("🗑️ Clear Results", use_container_width=True):
        clear_reconciliation_state()
        st.session_state.file_info = {"bank": None, "ledger": None}
        safe_rerun()
with col_btn3:
    if st.button("📥 Download Last Report", use_container_width=True, disabled=(st.session_state.output_bytes is None)):
        if st.session_state.output_bytes:
            st.download_button(
                label="Download",
                data=st.session_state.output_bytes,
                file_name=st.session_state.output_filename or "bank_reconciliation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_last"
            )

if run_recon:
    if not bank_file or not ledger_file:
        st.error("Please upload both Bank Statement and Ledger files")
        st.stop()
    with st.spinner("Processing..."):
        try:
            bank_df, bank_opening_auto, bank_closing_auto, _ = load_bank_statement(bank_file)
            ledger_df, ledger_closing, _ = load_ledger(ledger_file)
            opening_balance = st.session_state.opening_balance_manual if st.session_state.opening_balance_manual != 0 else (bank_opening_auto or 0)
            match_results = match_transactions(bank_df, ledger_df)
            working_paper = build_working_paper(match_results)
            recon_statement = build_recon_statement(
                opening_balance,
                bank_closing_auto,
                ledger_closing,
                match_results
            )
            output_bytes = export_to_excel(working_paper, recon_statement, bank_df, ledger_df, match_results)
            st.session_state.bank_df = bank_df
            st.session_state.ledger_df = ledger_df
            st.session_state.match_results = match_results
            st.session_state.working_paper = working_paper
            st.session_state.recon_statement = recon_statement
            st.session_state.output_bytes = output_bytes
            st.session_state.output_filename = f"bank_reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.session_state.reconciliation_done = True
            st.success(f"✅ Reconciliation complete! {len(match_results['matches'])} transactions matched.")
            safe_rerun()
        except Exception as e:
            st.error(f"Error: {e}")
            import traceback
            st.code(traceback.format_exc())

if st.session_state.reconciliation_done:
    st.markdown("---")
    st.info("💾 **Results from last reconciliation are shown below.** Upload new files and re-run to update.")
    match_results = st.session_state.match_results
    working_paper = st.session_state.working_paper
    recon_statement = st.session_state.recon_statement

    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
    with metric_col1:
        st.metric("Matched Transactions", len(match_results['matches']))
    with metric_col2:
        st.metric("Unmatched - Ledger", len(match_results['unmatched_ledger_credits']) + len(match_results['unmatched_ledger_debits']))
    with metric_col3:
        st.metric("Unmatched - Bank", len(match_results['unmatched_bank_credits']) + len(match_results['unmatched_bank_debits']))
    with metric_col4:
        diff = recon_statement['difference']
        st.metric("Balance Difference", f"{diff:,.2f}", delta="Zero" if abs(diff) < 0.01 else "Check")

    st.markdown("---")
    st.subheader("📋 Working Paper Preview")
    display_cols = ['SECTION', 'LEDGER_DATE', 'LEDGER_DESC', 'LEDGER_REF', 'LEDGER_AMOUNT', 
                    'MATCH_STATUS', 'BANK_AMOUNT', 'BANK_DATE', 'BANK_REF', 'BANK_DESC']
    wp_preview = working_paper[display_cols].head(30).copy()
    wp_preview['LEDGER_DATE'] = wp_preview['LEDGER_DATE'].apply(lambda x: format_date(x) if pd.notna(x) else '')
    wp_preview['BANK_DATE'] = wp_preview['BANK_DATE'].apply(lambda x: format_date(x) if pd.notna(x) else '')
    wp_preview['LEDGER_AMOUNT'] = wp_preview['LEDGER_AMOUNT'].apply(lambda x: f"{float(x):,.2f}" if pd.notna(x) and x != '' else '')
    wp_preview['BANK_AMOUNT'] = wp_preview['BANK_AMOUNT'].apply(lambda x: f"{float(x):,.2f}" if pd.notna(x) and x != '' else '')
    st.dataframe(wp_preview, use_container_width=True)

    st.markdown("---")
    st.subheader("📄 Clean Bank Reconciliation")
    recon_preview = [
        {"Reconciliation Step": "Opening Bank Balance", "Amount": f"{recon_statement['opening_balance']:,.2f}"},
        {"Reconciliation Step": "Closing Balance per Bank Statement", "Amount": f"{recon_statement['bank_closing_balance']:,.2f}"},
    ]
    for _, item in recon_statement['recon_items'].iterrows():
        recon_preview.append({
            "Reconciliation Step": item['description'],
            "Amount": f"{item['adjustment']:,.2f}"
        })
    recon_preview += [
        {"Reconciliation Step": "TOTAL RECONCILING ADJUSTMENTS", "Amount": f"{recon_statement['total_adjustment']:,.2f}"},
        {"Reconciliation Step": "ADJUSTED BANK BALANCE", "Amount": f"{recon_statement['adjusted_balance']:,.2f}"},
        {"Reconciliation Step": "LEDGER CLOSING BALANCE", "Amount": f"{recon_statement['ledger_balance']:,.2f}"},
        {"Reconciliation Step": "UNRECONCILED DIFFERENCE", "Amount": f"{recon_statement['difference']:,.2f}"},
    ]
    st.dataframe(pd.DataFrame(recon_preview), use_container_width=True)

    st.subheader("🔎 Why the Difference Exists")
    diagnostic = pd.DataFrame([
        {"Diagnostic": "Opening Bank Balance", "Amount": recon_statement['opening_balance']},
        {"Diagnostic": "Bank Movement During Period", "Amount": recon_statement['bank_movement']},
        {"Diagnostic": "Ledger Movement Using Same Opening Basis", "Amount": recon_statement['ledger_movement']},
        {"Diagnostic": "Movement Difference", "Amount": recon_statement['movement_difference']},
        {"Diagnostic": "Final Unreconciled Difference", "Amount": recon_statement['difference']},
    ])
    st.dataframe(diagnostic, use_container_width=True)

    if abs(recon_statement['difference']) < 0.01:
        st.success("✅ Reconciled: the adjusted bank balance agrees to the ledger.")
    else:
        st.warning(
            "⚠️ Investigation required. The system is deliberately not forcing the balance to zero. "
            "Review the unmatched transactions and opening balance difference before posting adjustments."
        )

st.markdown("")
logout_c1, logout_c2, logout_c3 = st.columns([1, 1, 1])
with logout_c2:
    if st.button("Logout", use_container_width=True):
        logout()
